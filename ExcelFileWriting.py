# -*- coding: utf-8 -*-
"""
Created on Fri Aug 18 14:11:23 2017

@author: BALASUBRAMANIAM
"""
from openpyxl import load_workbook
import random
filePath="AnnualReport2018.xlsx"
fileRef= load_workbook(filePath,read_only=False)
sheetNames=fileRef.get_sheet_names()
print(sheetNames)
sheet = fileRef.get_sheet_by_name('January_2018')
for row in range(51,100):
    for col in range(1,6):
        sheet.cell(column=col,row=row, 
                   value="%d" % random.randint(1,10000))
        
fileRef.save(filePath)







