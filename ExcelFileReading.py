# -*- coding: utf-8 -*-
"""
Created on Fri Aug 18 14:33:57 2017

@author: BALASUBRAMANIAM
"""
from openpyxl import load_workbook
from openpyxl.chart import (AreaChart,Series,BarChart3D,PieChart,Reference)
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Color, Fill
filePath="GDP.xlsx"
wb=load_workbook(filePath,read_only=False,data_only=True)

sheetNames=wb.get_sheet_names()
print(sheetNames)

sheet=wb.get_sheet_by_name('GDP')

for row in range(6,50):
    for col in range(1,6):
        print(sheet.cell(row=row,column=col).value)

        
chart=PieChart()
chart.title="World GDP Ranking"
chart.style=10
#chart.x_axis.title="Country Code"
#chart.y_axis.title="Rank"

#props = GraphicalProperties(solidFill="999999")
#chart.plot_area.graphicalProperties = props


xdata=Reference(sheet,min_col=1,min_row=6,max_row=50)
ydata=Reference(sheet,min_col=2,min_row=6,max_row=50)
chart.add_data(ydata,titles_from_data=True)
chart.set_categories(xdata)
sheet.add_chart(chart,"G10")
wb.save(filePath)

